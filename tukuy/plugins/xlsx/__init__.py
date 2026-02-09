"""Excel (XLSX) processing plugin.

Provides reading, writing, and conversion of Excel spreadsheets.

Requires ``openpyxl`` (optional, fails gracefully at runtime).
"""

import csv
import io
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from ...base import ChainableTransformer
from ...types import TransformContext
from ...plugins.base import TransformerPlugin
from ...safety import check_read_path, check_write_path
from ...skill import skill, ConfigParam, ConfigScope, RiskLevel


# ── Transformers ──────────────────────────────────────────────────────────

class XlsxToJsonTransformer(ChainableTransformer[str, str]):
    """Convert an XLSX file to a JSON string.

    Input is a file path. Reads the first (or specified) sheet and returns
    a JSON array of row-objects keyed by header names.
    Requires ``openpyxl``.
    """

    def __init__(self, name: str, sheet: Optional[str] = None, header_row: int = 1):
        super().__init__(name)
        self.sheet = sheet
        self.header_row = header_row

    def validate(self, value: str) -> bool:
        return isinstance(value, str)

    def _transform(self, value: str, context: Optional[TransformContext] = None) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return json.dumps({"error": "openpyxl is required. Install with: pip install openpyxl"})

        wb = load_workbook(value, read_only=True, data_only=True)
        ws = wb[self.sheet] if self.sheet and self.sheet in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return json.dumps([])

        header_idx = self.header_row - 1
        if header_idx >= len(rows):
            return json.dumps([])

        headers = [str(h or f"col_{i}") for i, h in enumerate(rows[header_idx])]
        data = []
        for row in rows[header_idx + 1:]:
            row_dict = {}
            for i, header in enumerate(headers):
                cell = row[i] if i < len(row) else None
                # Convert to JSON-serializable types
                if cell is None:
                    row_dict[header] = None
                elif isinstance(cell, (int, float, bool)):
                    row_dict[header] = cell
                else:
                    row_dict[header] = str(cell)
            data.append(row_dict)

        return json.dumps(data, ensure_ascii=False, default=str)


class JsonToXlsxTransformer(ChainableTransformer[dict, dict]):
    """Write JSON data to an XLSX file.

    Expects input as ``{"data": [...], "output": "file.xlsx", "sheet": "Sheet1"}``.
    ``data`` is a list of dicts (rows keyed by headers).
    Requires ``openpyxl``.
    """

    def validate(self, value: Any) -> bool:
        return isinstance(value, dict) and "data" in value

    def _transform(self, value: Any, context: Optional[TransformContext] = None) -> dict:
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"error": "openpyxl is required. Install with: pip install openpyxl"}

        data = value["data"]
        if isinstance(data, str):
            data = json.loads(data)

        output = value.get("output", "output.xlsx")
        sheet_name = value.get("sheet", "Sheet1")

        if not isinstance(data, list) or not data:
            return {"error": "data must be a non-empty list of dicts"}

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Headers from first row keys
        headers = list(data[0].keys())
        ws.append(headers)

        # Data rows
        for row in data:
            ws.append([row.get(h) for h in headers])

        wb.save(output)
        return {
            "output": output,
            "sheet": sheet_name,
            "rows_written": len(data),
            "columns": headers,
            "success": True,
        }


class XlsxToCsvTransformer(ChainableTransformer[str, str]):
    """Convert an XLSX sheet to CSV string.

    Input is a file path. Returns CSV text.
    Requires ``openpyxl``.
    """

    def __init__(self, name: str, sheet: Optional[str] = None):
        super().__init__(name)
        self.sheet = sheet

    def validate(self, value: str) -> bool:
        return isinstance(value, str)

    def _transform(self, value: str, context: Optional[TransformContext] = None) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "[error] openpyxl is required. Install with: pip install openpyxl"

        wb = load_workbook(value, read_only=True, data_only=True)
        ws = wb[self.sheet] if self.sheet and self.sheet in wb.sheetnames else wb.active

        out = io.StringIO()
        writer = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(c) if c is not None else "" for c in row])
        wb.close()

        return out.getvalue().strip()


# ── Skills ────────────────────────────────────────────────────────────────

@skill(
    name="xlsx_read",
    description="Read an Excel file and return sheet data, cell values, and metadata.",
    category="xlsx",
    tags=["xlsx", "excel", "spreadsheet"],
    idempotent=True,
    requires_filesystem=True,
    required_imports=["openpyxl"],
    display_name="Read Excel",
    icon="table",
    risk_level=RiskLevel.SAFE,
    group="Excel",
    config_params=[
        ConfigParam(
            name="max_rows_per_sheet",
            display_name="Max Rows per Sheet",
            description="Maximum rows to read per sheet.",
            type="number",
            default=1000,
            min=1,
            max=1000000,
        ),
    ],
)
def xlsx_read(
    path: str,
    sheet: Optional[str] = None,
    max_rows: int = 1000,
) -> dict:
    """Read an Excel file and return its content."""
    path = check_read_path(path)
    p = Path(path)
    if not p.exists():
        return {"path": path, "error": "File not found", "exists": False}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl is required. Install with: pip install openpyxl"}

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet] if sheet and sheet in sheet_names else wb.active

    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h or f"col_{j}") for j, h in enumerate(row)]
            continue
        if i > max_rows:
            break
        row_dict = {}
        for j, header in enumerate(headers):
            cell = row[j] if j < len(row) else None
            if cell is None:
                row_dict[header] = None
            elif isinstance(cell, (int, float, bool)):
                row_dict[header] = cell
            else:
                row_dict[header] = str(cell)
        rows.append(row_dict)

    dimensions = ws.dimensions
    wb.close()

    return {
        "path": path,
        "exists": True,
        "sheet": ws.title,
        "sheets": sheet_names,
        "headers": headers,
        "rows": rows,
        "row_count": len(rows),
        "dimensions": dimensions,
    }


@skill(
    name="xlsx_write",
    description="Write data to an Excel file. Creates a new file or overwrites existing.",
    category="xlsx",
    tags=["xlsx", "excel", "spreadsheet"],
    side_effects=True,
    requires_filesystem=True,
    required_imports=["openpyxl"],
    display_name="Write Excel",
    icon="table",
    risk_level=RiskLevel.MODERATE,
    group="Excel",
)
def xlsx_write(
    path: str,
    data: list,
    sheet: str = "Sheet1",
    headers: Optional[list] = None,
) -> dict:
    """Write data to an Excel file."""
    path = check_write_path(path)
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"error": "openpyxl is required. Install with: pip install openpyxl"}

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    if not data:
        wb.save(path)
        return {"path": path, "rows_written": 0, "success": True}

    # Determine headers
    if headers:
        ws.append(headers)
    elif isinstance(data[0], dict):
        headers = list(data[0].keys())
        ws.append(headers)

    # Write rows
    for row in data:
        if isinstance(row, dict):
            ws.append([row.get(h) for h in (headers or row.keys())])
        elif isinstance(row, (list, tuple)):
            ws.append(list(row))
        else:
            ws.append([row])

    wb.save(path)
    return {
        "path": path,
        "sheet": sheet,
        "rows_written": len(data),
        "columns": headers or [],
        "success": True,
    }


@skill(
    name="xlsx_sheets",
    description="List all sheet names and dimensions in an Excel file.",
    category="xlsx",
    tags=["xlsx", "excel", "spreadsheet"],
    idempotent=True,
    requires_filesystem=True,
    required_imports=["openpyxl"],
    display_name="List Sheets",
    icon="list",
    risk_level=RiskLevel.SAFE,
    group="Excel",
)
def xlsx_sheets(path: str) -> dict:
    """List sheets in an Excel file."""
    path = check_read_path(path)
    p = Path(path)
    if not p.exists():
        return {"path": path, "error": "File not found", "exists": False}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl is required. Install with: pip install openpyxl"}

    wb = load_workbook(path, read_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "dimensions": ws.dimensions,
        })
    wb.close()

    return {
        "path": path,
        "exists": True,
        "sheets": sheets,
        "count": len(sheets),
    }


class XlsxPlugin(TransformerPlugin):
    """Plugin providing Excel spreadsheet processing."""

    def __init__(self):
        super().__init__("xlsx")

    @property
    def transformers(self) -> Dict[str, callable]:
        return {
            "xlsx_to_json": lambda params: XlsxToJsonTransformer(
                "xlsx_to_json",
                sheet=params.get("sheet"),
                header_row=params.get("header_row", 1),
            ),
            "json_to_xlsx": lambda _: JsonToXlsxTransformer("json_to_xlsx"),
            "xlsx_to_csv": lambda params: XlsxToCsvTransformer(
                "xlsx_to_csv",
                sheet=params.get("sheet"),
            ),
        }

    @property
    def skills(self) -> Dict[str, Any]:
        return {
            "xlsx_read": xlsx_read.__skill__,
            "xlsx_write": xlsx_write.__skill__,
            "xlsx_sheets": xlsx_sheets.__skill__,
        }

    @property
    def manifest(self):
        from ...manifest import PluginManifest, PluginRequirements
        return PluginManifest(
            name="xlsx",
            display_name="Excel",
            description="Read, write, and inspect Excel spreadsheet files.",
            icon="table",
            group="Documents",
            requires=PluginRequirements(filesystem=True, imports=["openpyxl"]),
        )
